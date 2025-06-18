import pyodbc
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import os
import streamlit as st

# SQL connection string
edlive_conn_str = (
    'DRIVER={ODBC Driver 17 for SQL Server};'
    'SERVER=nav18db;'
    'DATABASE=EDLIVE;'
    'UID=barcode1;'
    'PWD=barcode@1433;'
    'Trusted_Connection=no;'
)

# SQL query (unchanged)
sql_query = """ 
WITH BaseData AS (
    SELECT 
        h.No_ AS [Transfer No],
        h.[Transfer-from Code],
        h.[Transfer-to Code],
        h.[Created by WS],
        'Open' AS Status,
        h.[Posting Date],
        h.[External Document No_],
        h.[External Document No_ 2],
        l.[Document No_] AS [Transfer Document No],
        l.[Outstanding Quantity] AS [Transfer Quantity],
        re.Quantity AS [Reserved Quantity],
        l.[Item No_],
        l.[Line No_],
        ISNULL(ile.RemainingQuantity, 0) AS [Initial Stock],
        re.[Source ID],
        re.[Source Ref_ No_],
        CASE 
            WHEN re.Positive = 0 THEN 'From ILE (Available Stock)'
            WHEN re.Positive = 1 THEN 'From Pending Transfer Order'
            ELSE 'YetReserved'
        END AS [Reservation Source]
    FROM 
        [EDLIVE].[dbo].[Eastern Decorator Sdn_ Bhd_$Transfer Header] h
    INNER JOIN 
        [EDLIVE].[dbo].[Eastern Decorator Sdn_ Bhd_$Transfer Line] l
        ON h.No_ = l.[Document No_]
    LEFT JOIN (
        SELECT 
            [Item No_],  
            [Location Code],  
            SUM([Remaining Quantity]) AS RemainingQuantity 
        FROM 
            [EDLIVE].[dbo].[Eastern Decorator Sdn_ Bhd_$Item Ledger Entry]
        WHERE 
            [Remaining Quantity] > 0
            AND [Entry Type] in('4','2','1','0') 
        GROUP BY 
            [Item No_], 
            [Location Code]
    ) ile
        ON l.[Item No_] = ile.[Item No_]
        AND h.[Transfer-from Code] = ile.[Location Code]
    LEFT JOIN 
        [EDLIVE].[dbo].[Eastern Decorator Sdn_ Bhd_$Reservation Entry] re
        ON h.No_ = re.[Source ID]
        AND l.[Line No_] = re.[Source Ref_ No_]
        AND re.[Source ID] IS NOT NULL
        AND LTRIM(RTRIM(re.[Source ID])) <> ''
    WHERE 
        h.[Created By User ID] = 'EADECO\\CAAPI' 
        AND h.Status = 0
        AND (
            h.[External Document No_ 2] IS NULL 
            OR h.[External Document No_ 2] NOT LIKE 'E\\_%' ESCAPE '\\'
        )
)

SELECT 
    [Transfer No],
    [Transfer-from Code],
    [Transfer-to Code],
    [Created by WS],
    Status,
    [Posting Date],
    [External Document No_],
    [External Document No_ 2],
    [Transfer Document No],
    [Item No_],
    [Line No_],
    [Transfer Quantity],
    [Reserved Quantity],
    CASE 
        WHEN SUM(
            CASE 
                WHEN [Reserved Quantity] < 0 THEN ABS([Reserved Quantity]) 
                ELSE ISNULL([Reserved Quantity], 0) 
            END
        ) OVER (
            PARTITION BY [Transfer Document No], [Item No_], [Line No_]
            ORDER BY [Posting Date], [Source ID], [Source Ref_ No_]
            ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
        ) IS NOT NULL
        THEN 
            [Transfer Quantity] - 
            SUM(
                CASE 
                    WHEN [Reserved Quantity] < 0 THEN ABS([Reserved Quantity]) 
                    ELSE ISNULL([Reserved Quantity], 0) 
                END
            ) OVER (
                PARTITION BY [Transfer Document No], [Item No_], [Line No_]
                ORDER BY [Posting Date], [Source ID], [Source Ref_ No_]
                ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
            )
        ELSE [Transfer Quantity]
    END AS [Unreserved Quantity],
    [Initial Stock],
    [Initial Stock] + 
        SUM(ISNULL([Reserved Quantity], 0)) OVER (
            PARTITION BY [Item No_], [Transfer-from Code]
            ORDER BY [Posting Date], [Transfer Document No], [Line No_]
            ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
        ) AS [Available Quantity After Reservation],
    CASE
        WHEN 
            ([Initial Stock] + 
            SUM(ISNULL([Reserved Quantity], 0)) OVER (
                PARTITION BY [Item No_], [Transfer-from Code]
                ORDER BY [Posting Date], [Transfer Document No], [Line No_]
                ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
            )) = 0
        THEN 
            CASE 
                WHEN SUM(
                    CASE 
                        WHEN [Reserved Quantity] < 0 THEN ABS([Reserved Quantity]) 
                        ELSE ISNULL([Reserved Quantity], 0) 
                    END
                ) OVER (
                    PARTITION BY [Transfer Document No], [Item No_], [Line No_]
                    ORDER BY [Posting Date], [Source ID], [Source Ref_ No_]
                    ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
                ) IS NOT NULL
                THEN 
                    [Transfer Quantity] -
                    SUM(
                        CASE 
                            WHEN [Reserved Quantity] < 0 THEN ABS([Reserved Quantity]) 
                            ELSE ISNULL([Reserved Quantity], 0) 
                        END
                    ) OVER (
                        PARTITION BY [Transfer Document No], [Item No_], [Line No_]
                        ORDER BY [Posting Date], [Source ID], [Source Ref_ No_]
                        ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
                    )
                ELSE [Transfer Quantity]
            END
        ELSE 0
    END AS [BalanceToReserved],
    [Source ID],
    [Source Ref_ No_],
    [Reservation Source]
FROM BaseData
ORDER BY 
    [Posting Date] DESC,
    [Item No_],
    [Line No_];
"""

def get_data():
    try:
        conn = pyodbc.connect(edlive_conn_str)
        df = pd.read_sql(sql_query, conn)
        return df
    except Exception as e:
        st.error(f"❌ Database error: {str(e)}")
        return None
    finally:
        if 'conn' in locals():
            conn.close()

def format_excel_report(file_path):
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.1
            ws.column_dimensions[column_letter].width = adjusted_width

        ws.freeze_panes = 'B2'
        wb.save(file_path)
        return True
    except Exception as e:
        st.error(f"❌ Excel formatting error: {str(e)}")
        return False

def generate_excel_report(df):
    try:
        output_path = r"Z:\\CavsNavErrors\\TOYetReservedNoInventory"
        os.makedirs(output_path, exist_ok=True)

        filename = f"Transfer_Order_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_file = os.path.join(output_path, filename)

        export_df = df[
            (df['Reservation Source'] == 'YetReserved') & 
            (df['BalanceToReserved'] > 0)
        ].copy()

        # Add ILEPostingdate column as first day of Posting Date's month
        export_df['ILEPostingdate'] = pd.to_datetime(export_df['Posting Date']).dt.to_period('M').dt.to_timestamp().dt.strftime('%Y-%m-%d')

        # Group by ILEPostingdate + External Doc + Item + Location
        grouped = export_df.groupby(
            ['ILEPostingdate', 'Item No_', 'Transfer-from Code', 'External Document No_', 'External Document No_ 2'],
            as_index=False
        )['BalanceToReserved'].sum()

        # Rename for output
        grouped = grouped.rename(columns={
            'Transfer-from Code': 'Location Code'
        })

        # Reorder columns
        grouped = grouped[[
            'ILEPostingdate', 'Item No_', 'Location Code', 'External Document No_', 'External Document No_ 2', 'BalanceToReserved'
        ]]

        grouped.to_excel(output_file, index=False)

        if format_excel_report(output_file):
            return output_file
        return None
    except Exception as e:
        st.error(f"❌ Report generation error: {str(e)}")
        return None

def main():
    st.set_page_config(layout="wide")

    st.markdown("""
        <style>
        html, body, [class*="css"] {
            font-size: 0.9rem !important;
        }
        h1, h2 {
            margin-top: -20px !important;
            margin-bottom: 0.5rem;
        }
        .stDataFrame div[data-testid="stMarkdownContainer"],
        .stDataFrame table {
            font-size: 0.85rem !important;
        }
        </style>
    """, unsafe_allow_html=True)

    st.title("📋 Transfer Orders Report")

    if 'df' not in st.session_state:
        st.session_state.df = None
    if 'report_path' not in st.session_state:
        st.session_state.report_path = None

    with st.sidebar:
        st.header("🔧 Options")
        if st.button("📥 Load Transfer Orders"):
            with st.spinner("Fetching data..."):
                st.session_state.df = get_data()
                if st.session_state.df is not None:
                    filtered_df = st.session_state.df[
                        (st.session_state.df['Reservation Source'] == 'YetReserved') &
                        (st.session_state.df['BalanceToReserved'] > 0)
                    ].copy()

                    filtered_df['Posting Date'] = pd.to_datetime(filtered_df['Posting Date'])
                    grouped_df = filtered_df.groupby(
                        ['Transfer No', 'Item No_', 'Transfer-from Code', 'Transfer-to Code'],
                        as_index=False
                    ).agg({
                        'Posting Date': 'min',
                        'External Document No_': 'first',
                        'External Document No_ 2': 'first',
                        'Transfer Quantity': 'sum',
                        'BalanceToReserved': 'sum'
                    })

                    grouped_df['Posting Date'] = grouped_df['Posting Date'].dt.strftime('%Y-%m-%d')
                    st.session_state.filtered_df = grouped_df
                    st.success("✅ Data loaded and grouped successfully!")

        if st.button("📤 Generate Excel Report"):
            if st.session_state.df is None:
                st.warning("⚠️ Please load data first using 'Load Transfer Orders'")
            else:
                with st.spinner("Generating Excel report..."):
                    path = generate_excel_report(st.session_state.df)
                    if path:
                        st.session_state.report_path = path
                        st.success("✅ Excel report generated!")

    if 'filtered_df' in st.session_state and st.session_state.filtered_df is not None:
        st.subheader("📊 Open Transfer Orders")
        st.dataframe(st.session_state.filtered_df, use_container_width=True)

    if st.session_state.report_path:
        with open(st.session_state.report_path, "rb") as f:
            st.download_button(
                label="⬇️ Download Excel Report",
                data=f,
                file_name=os.path.basename(st.session_state.report_path),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

if __name__ == "__main__":
    main()
