import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
import pyodbc
from datetime import datetime, timedelta
import os
import io
import logging

# Your DB config and logging setup here (same as your code)...
bcs = {
    'server': 'caappsdb,1435',
    'database': 'BCSSoft_ConAppSys',
    'username': 'Deepak',
    'password': 'Deepak@321',
    'driver': 'ODBC Driver 17 for SQL Server',
    'timeout': 30
}

nav = {
    'server': 'nav18db',
    'database': 'EDLIVE',
    'username': 'barcode1',
    'password': 'barcode@1433',
    'driver': 'ODBC Driver 17 for SQL Server',
    'timeout': 30
}

log_file = os.path.join(os.environ['USERPROFILE'], 'Documents', 'cr_report_scheduler.log')
logging.basicConfig(
    filename=log_file,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def get_bcs_connection():
    conn_str = f"DRIVER={{{bcs['driver']}}};SERVER={bcs['server']};DATABASE={bcs['database']};UID={bcs['username']};PWD={bcs['password']};Timeout={bcs['timeout']}"
    return pyodbc.connect(conn_str)

def get_nav_connection():
    conn_str = f"DRIVER={{{nav['driver']}}};SERVER={nav['server']};DATABASE={nav['database']};UID={nav['username']};PWD={nav['password']};Timeout={nav['timeout']}"
    return pyodbc.connect(conn_str)

def generate_cr_report_df():
    three_months_ago = datetime.now() - timedelta(days=90)
    three_months_ago_str = three_months_ago.strftime('%Y-%m-%d')

    sql = f"""
    SELECT 
        h.CustRtnHeaderId,
        h.CRNo,
        h.ClosedDt,
        d.CustRtnDetailId,
        d.CustRtnTypeCode,
        d.CreatedDt,
        log.LogTypeCode,
        log.LogStsCode,
        log.LogMsg
    FROM 
        [BCSSoft_ConAppSys].[dbo].[tbCustRtnHeader] h
    INNER JOIN 
        [BCSSoft_ConAppSys].[dbo].[tbCustRtnDetail] d 
        ON h.CustRtnHeaderId = d.CustRtnHeaderId
    INNER JOIN 
        [BCSSoft_ConAppSys].[dbo].[tbIntgNavLog] log 
        ON log.RefHdrId = d.CustRtnHeaderId 
        AND log.RefDtlId = d.CustRtnDetailId
    WHERE 
        d.CustRtnTypeCode IN ('SalesReconcile', 'Refund', 'Exchange')
        AND log.LogStsCode = 'E'
        AND d.CreatedDt >= '{three_months_ago_str}'
    ORDER BY 
        d.CreatedDt DESC
    """

    with get_bcs_connection() as conn:
        df = pd.read_sql(sql, conn)

    if df.empty:
        return pd.DataFrame()  # no data

    def fix_crno(crno):
        if pd.isna(crno):
            return ""
        try:
            part = crno.split('CR_')[-1].replace('_', '')
            if len(part) > 10 and part[8:10] == '20':
                part = part[:8] + part[10:]
            return part
        except Exception:
            return crno

    df['Document No_'] = df['CRNo'].apply(fix_crno)
    doc_nos = df['Document No_'].unique().tolist()

    def chunks(lst, n):
        for i in range(0, len(lst), n):
            yield lst[i:i + n]

    nav_docs_set = set()
    with get_nav_connection() as nav_conn:
        for batch in chunks(doc_nos, 1000):
            in_clause = ",".join(f"'{doc}'" for doc in batch)
            nav_sql = f"""
            SELECT DISTINCT [Document No_]
            FROM [EDLIVE].[dbo].[Eastern Decorator Sdn_ Bhd_$Item Ledger Entry]
            WHERE [Entry Type] = '1'
              AND [Document No_] IN ({in_clause})
            """
            nav_df = pd.read_sql(nav_sql, nav_conn)
            nav_docs_set.update(nav_df['Document No_'].tolist())

    df['IsPostedILE?'] = df['Document No_'].apply(lambda x: 'OK' if x in nav_docs_set else 'NotOK')

    cols = list(df.columns)
    cols.remove('Document No_')
    cols.remove('IsPostedILE?')
    cols.insert(6, 'Document No_')
    cols.insert(7, 'IsPostedILE?')
    df = df[cols]

    return df

def df_to_excel_bytes(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "CR Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for col_num, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if df.columns[c_idx - 1] == 'IsPostedILE?':
                cell.font = Font(color="008000" if value == 'OK' else "FF0000")

    for col_cells in ws.columns:
        max_length = 0
        column_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[column_letter].width = max_length + 2

    with io.BytesIO() as output:
        wb.save(output)
        data = output.getvalue()
    return data

# --- Streamlit App ---

st.title("CR Report Viewer & Exporter")

if st.button("Generate CR Report"):
    with st.spinner("Fetching data and generating report..."):
        try:
            df_report = generate_cr_report_df()
            if df_report.empty:
                st.warning("No data found for the last 3 months with LogStsCode='E'.")
            else:
                st.success(f"Report generated! {len(df_report)} rows retrieved.")
                st.dataframe(df_report)

                excel_bytes = df_to_excel_bytes(df_report)
                st.download_button(
                    label="Download Excel Report",
                    data=excel_bytes,
                    file_name="CR_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception as e:
            st.error(f"Error generating report: {e}")
