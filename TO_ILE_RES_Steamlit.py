import pyodbc
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import os
import streamlit as st

# Database connection string
edlive_conn_str = (
    'DRIVER={ODBC Driver 17 for SQL Server};'
    'SERVER=nav18db;'
    'DATABASE=EDLIVE;'
    'UID=barcode1;'
    'PWD=barcode@1433;'
    'Trusted_Connection=no;'
)

# SQL Query (same as your original)
sql_query = """
WITH BaseData AS (
    SELECT 
        h.No_ AS [Transfer No],
        h.[Transfer-from Code],
        h.[Transfer-to Code],
        h.[Created by WS],
        'Open' AS Status,
        h.[Posting Date],
        h.[External Document No_],
        h.[External Document No_ 2],
        
        l.[Document No_] AS [Transfer Document No],
        l.[Outstanding Quantity] AS [Transfer Quantity],
        re.Quantity AS [Reserved Quantity],
        l.[Item No_],
        l.[Line No_],
        
        ISNULL(ile.RemainingQuantity, 0) AS [Initial Stock],
        re.[Source ID],
        re.[Source Ref_ No_],
        CASE 
            WHEN re.Positive = 0 THEN 'From ILE (Available Stock)'
            WHEN re.Positive = 1 THEN 'From Pending Transfer Order'
            ELSE 'YetReserved'
        END AS [Reservation Source]
    FROM 
        [EDLIVE].[dbo].[Eastern Decorator Sdn_ Bhd_$Transfer Header] h
    INNER JOIN 
        [EDLIVE].[dbo].[Eastern Decorator Sdn_ Bhd_$Transfer Line] l
        ON h.No_ = l.[Document No_]
    LEFT JOIN (
        SELECT 
            [Item No_],  
            [Location Code],  
            SUM([Remaining Quantity]) AS RemainingQuantity 
        FROM 
            [EDLIVE].[dbo].[Eastern Decorator Sdn_ Bhd_$Item Ledger Entry]
        WHERE 
            [Remaining Quantity] > 0
            AND [Entry Type] in('4','2','1','0') 
        GROUP BY 
            [Item No_], 
            [Location Code]
    ) ile
        ON l.[Item No_] = ile.[Item No_]
        AND h.[Transfer-from Code] = ile.[Location Code]
    LEFT JOIN 
        [EDLIVE].[dbo].[Eastern Decorator Sdn_ Bhd_$Reservation Entry] re
        ON h.No_ = re.[Source ID]
        AND l.[Line No_] = re.[Source Ref_ No_]
        AND re.[Source ID] IS NOT NULL
        AND LTRIM(RTRIM(re.[Source ID])) <> ''
    WHERE 
        h.[Created By User ID] = 'EADECO\\CAAPI' 
        AND h.Status = 0
        AND (
            h.[External Document No_ 2] IS NULL 
            OR h.[External Document No_ 2] NOT LIKE 'E\\_%' ESCAPE '\\'
        )
)

SELECT 
    [Transfer No],
    [Transfer-from Code],
    [Transfer-to Code],
    [Created by WS],
    Status,
    [Posting Date],
    [External Document No_],
    [External Document No_ 2],
    [Transfer Document No],
    [Item No_],
    [Line No_],
    [Transfer Quantity],
    [Reserved Quantity],
    CASE 
        WHEN SUM(
            CASE 
                WHEN [Reserved Quantity] < 0 THEN ABS([Reserved Quantity]) 
                ELSE ISNULL([Reserved Quantity], 0) 
            END
        ) OVER (
            PARTITION BY [Transfer Document No], [Item No_], [Line No_]
            ORDER BY [Posting Date], [Source ID], [Source Ref_ No_]
            ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
        ) IS NOT NULL
        THEN 
            [Transfer Quantity] - 
            SUM(
                CASE 
                    WHEN [Reserved Quantity] < 0 THEN ABS([Reserved Quantity]) 
                    ELSE ISNULL([Reserved Quantity], 0) 
                END
            ) OVER (
                PARTITION BY [Transfer Document No], [Item No_], [Line No_]
                ORDER BY [Posting Date], [Source ID], [Source Ref_ No_]
                ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
            )
        ELSE [Transfer Quantity]
    END AS [Unreserved Quantity],
    [Initial Stock],
    [Initial Stock] + 
        SUM(ISNULL([Reserved Quantity], 0)) OVER (
            PARTITION BY [Item No_], [Transfer-from Code]
            ORDER BY [Posting Date], [Transfer Document No], [Line No_]
            ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
        ) AS [Available Quantity After Reservation],
    CASE
        WHEN 
            ([Initial Stock] + 
            SUM(ISNULL([Reserved Quantity], 0)) OVER (
                PARTITION BY [Item No_], [Transfer-from Code]
                ORDER BY [Posting Date], [Transfer Document No], [Line No_]
                ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
            )) = 0
        THEN 
            CASE 
                WHEN SUM(
                    CASE 
                        WHEN [Reserved Quantity] < 0 THEN ABS([Reserved Quantity]) 
                        ELSE ISNULL([Reserved Quantity], 0) 
                    END
                ) OVER (
                    PARTITION BY [Transfer Document No], [Item No_], [Line No_]
                    ORDER BY [Posting Date], [Source ID], [Source Ref_ No_]
                    ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
                ) IS NOT NULL
                THEN 
                    [Transfer Quantity] -
                    SUM(
                        CASE 
                            WHEN [Reserved Quantity] < 0 THEN ABS([Reserved Quantity]) 
                            ELSE ISNULL([Reserved Quantity], 0) 
                        END
                    ) OVER (
                        PARTITION BY [Transfer Document No], [Item No_], [Line No_]
                        ORDER BY [Posting Date], [Source ID], [Source Ref_ No_]
                        ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
                    )
                ELSE [Transfer Quantity]
            END
        ELSE 0
    END AS [NeedAdjQty],
    [Source ID],
    [Source Ref_ No_],
    [Reservation Source]
FROM BaseData
ORDER BY 
    [Posting Date] DESC,
    [Item No_],
    [Line No_];
"""

def get_data():
    """Retrieve data from database"""
    try:
        conn = pyodbc.connect(edlive_conn_str)
        df = pd.read_sql(sql_query, conn)
        return df
    except Exception as e:
        st.error(f"Database error: {str(e)}")
        return None
    finally:
        if 'conn' in locals():
            conn.close()

def format_excel_report(file_path):
    """Apply formatting to the Excel report"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        ws.freeze_panes = 'A2'
        wb.save(file_path)
        return True
    except Exception as e:
        st.error(f"Error formatting Excel file: {str(e)}")
        return False

def generate_excel_report(df, report_type="full"):
    """Generate Excel report and return file path"""
    try:
        output_path = r"Z:\CavsNavErrors\TOYetReservedNoInventory"
        os.makedirs(output_path, exist_ok=True)
        
        if report_type == "full":
            filename = "Transfer_Report.xlsx"
            df_to_export = df
        else:
            filename = "Adjustment_Report.xlsx"
            df_to_export = df[df['NeedAdjQty'] > 0][[
                'Transfer-from Code', 
                'External Document No_',
                'External Document No_ 2',
                'Item No_',
                'NeedAdjQty'
            ]].rename(columns={'Transfer-from Code': 'Location Code'})
        
        output_file = os.path.join(output_path, filename)
        df_to_export.to_excel(output_file, index=False)
        
        if format_excel_report(output_file):
            return output_file
        return None
    except Exception as e:
        st.error(f"Error generating report: {str(e)}")
        return None

def main():
    st.title("Transfer Order Reports")
    st.write("Generate different reports for transfer orders")
    
    # Initialize session state for data
    if 'df' not in st.session_state:
        st.session_state.df = None
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("View Transfer Report"):
            with st.spinner("Loading data..."):
                st.session_state.df = get_data()
                if st.session_state.df is not None:
                    st.dataframe(st.session_state.df)
    
    with col2:
        if st.button("View NeedAdjQty Report"):
            if st.session_state.df is None:
                st.warning("Please load data first using 'View Transfer Report'")
            else:
                adj_df = st.session_state.df[st.session_state.df['NeedAdjQty'] > 0][[
                    'Transfer-from Code', 
                    'External Document No_',
                    'External Document No_ 2',
                    'Item No_',
                    'NeedAdjQty'
                ]].rename(columns={'Transfer-from Code': 'Location Code'})
                
                if not adj_df.empty:
                    st.dataframe(adj_df)
                else:
                    st.info("No items with NeedAdjQty > 0 found")
    
    with col3:
        if st.button("Generate Excel Reports"):
            if st.session_state.df is None:
                st.warning("Please load data first using 'View Transfer Report'")
            else:
                with st.spinner("Generating reports..."):
                    # Generate full report
                    full_report = generate_excel_report(st.session_state.df, "full")
                    # Generate adjustment report
                    adj_report = generate_excel_report(st.session_state.df, "adj")
                    
                    if full_report and adj_report:
                        st.success(f"Reports generated successfully at:\n- {full_report}\n- {adj_report}")
                    elif full_report:
                        st.success(f"Transfer report generated at: {full_report}")
                        st.warning("Adjustment report not generated (no items with NeedAdjQty > 0)")
                    else:
                        st.error("Failed to generate reports")

if __name__ == "__main__":
    main()