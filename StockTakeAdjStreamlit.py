import streamlit as st
import pandas as pd
import pyodbc
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Database configuration
CONFIG = {
    'server': 'caappsdb,1435',
    'database': 'BCSSoft_ConAppSys',
    'username': 'Deepak',
    'password': 'Deepak@321',
    'driver': 'ODBC Driver 17 for SQL Server'
}

# SQL query
sql_query = """
WITH DetailWithAdjustment AS (
    SELECT 
        H.CounterCode,
        H.CCName,
        H.ConfirmedDt,
        D.Location,
        D.SKU,
        D.QtyCounted,
        D.SystemQty,
        D.QtyAdjusted AS CCDetail_QtyAdjusted,
        A.Qty AS CCAdj_Qty,
        A.AdjDt,
        A.SourceFrom,
        CASE 
            WHEN A.CounterCode IS NOT NULL THEN 'Yes' 
            ELSE 'No' 
        END AS HasAdjustment
    FROM [dbo].[tbCCHeader] H
    JOIN [dbo].[tbCCDetail] D
        ON H.CCHeaderID = D.CCHeaderID
    LEFT JOIN [dbo].[tbInvAdj] A
        ON H.CounterCode = A.CounterCode 
        AND D.SKU = A.SKU
        AND A.SourceFrom = 'CCadj'
        AND A.AdjDt = EOMONTH(H.ConfirmedDt, -1)
    WHERE 
        H.CCTypeCode = 'FST' 
        AND H.ConfirmedDt >= DATEADD(month, -3, GETDATE())
)

SELECT 
    d.CounterCode,
    d.CCName,
    d.SKU,
    d.ConfirmedDt,
    SUM(d.SystemQty) AS TotalSystemQty,
    SUM(d.QtyCounted) AS TotalQtyCounted,
    d.CCAdj_Qty,
    d.AdjDt AS AdjustmentDate,
    d.HasAdjustment
FROM DetailWithAdjustment d
WHERE d.HasAdjustment = 'Yes'
GROUP BY 
    d.CounterCode,
    d.CCName,
    d.SKU,
    d.ConfirmedDt,
    d.CCDetail_QtyAdjusted,
    d.CCAdj_Qty,
    d.AdjDt,
    d.HasAdjustment
"""

def get_connection():
    conn_str = f"DRIVER={{{CONFIG['driver']}}};SERVER={CONFIG['server']};DATABASE={CONFIG['database']};UID={CONFIG['username']};PWD={CONFIG['password']}"
    return pyodbc.connect(conn_str)

def create_excel(df: pd.DataFrame) -> BytesIO:
    # Create Excel in-memory file with styling
    output = BytesIO()
    df.to_excel(output, index=False, sheet_name='Adjustments')
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active

    # Freeze header row
    ws.freeze_panes = "A2"

    # Style header row
    header_font = Font(color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Bold highlight columns
    highlight_columns = ['CCAdj_Qty', 'OnHandQty']
    for col in ws.iter_cols():
        if col[0].value in highlight_columns:
            for cell in col:
                cell.font = Font(bold=True)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = (max_length + 2) * 1.2

    styled_output = BytesIO()
    wb.save(styled_output)
    styled_output.seek(0)
    return styled_output

def main():
    st.title("Cycle Count Adjustments Report")

    try:
        with st.spinner("Loading data from database..."):
            conn = get_connection()
            df = pd.read_sql(sql_query, conn)
            conn.close()

        if df.empty:
            st.warning("No adjustment records found for the last 3 months.")
            return

        # Calculate OnHandQty and drop TotalSystemQty
        df['OnHandQty'] = df['TotalSystemQty'] - df['CCAdj_Qty']
        df.drop(columns=['TotalSystemQty'], inplace=True)

        # Reorder columns for display/export
        columns_order = [
            'CounterCode', 'CCName', 'SKU', 'ConfirmedDt',
            'CCAdj_Qty', 'OnHandQty', 'TotalQtyCounted',
            'AdjustmentDate', 'HasAdjustment'
        ]
        df = df[columns_order]

        # Show dataframe in Streamlit
        st.dataframe(df, use_container_width=True)

        # Button to download Excel
        excel_data = create_excel(df)
        st.download_button(
            label="📥 Download Excel Report",
            data=excel_data,
            file_name="CycleCount_Adjustments.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except pyodbc.Error as db_err:
        st.error(f"Database error: {db_err}")
    except Exception as e:
        st.error(f"Unexpected error: {e}")

if __name__ == "__main__":
    main()
